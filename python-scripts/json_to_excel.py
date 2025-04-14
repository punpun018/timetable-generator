import json
import pandas as pd
import sys
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Side, Border

def format_sheet(worksheet, dataframe) -> None:
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    row_fill_1 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    row_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                          top=Side(style="thin"), bottom=Side(style="thin"))

    for col_num, cell in enumerate(worksheet[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        col_letter = get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 20

    for row_num, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(dataframe) + 1), 2):
        fill = row_fill_1 if row_num % 2 == 0 else row_fill_2
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            cell.border = thin_border


def json_to_excel(json_data: dict) -> None:
    try:
        data = json_data

        num_days = len(data["timeTable"])
        num_slots = len(data["timeTable"][0][1])

        slots = [""] * num_slots
        for i in range(num_slots):
            slots[i] = ["\n".join(day[i][1]) for day in data["timeTable"]]

        dataframe1 = pd.DataFrame({
            "Day": [f"{i + 1}" for i in range(num_days)]
        })
        for i, slot in enumerate(slots):
            dataframe1[f"Slot {i + 1}"] = slot

        dataframe5 = pd.DataFrame({
            "Day": [f"{i + 1}" for i in range(num_days)]
        })
        for i in range(num_slots):
            dataframe5[f"Slot {i + 1}"] = [day[i][0] for day in data["timeTable"]]

        dataframe2 = pd.DataFrame({
            "Student": [student["rollNumber"] for student in data["timeTableSummary"]]
        })
        for day in range(len(data["timeTable"])):
            dataframe2[f"Day {day + 1}"] = [student["exams"][day] for student in data["timeTableSummary"]]

        dataframe3 = pd.DataFrame({
            "Student": [student["rollNumber"] for student in data["timeTableSummary"]]
        })
        for day in range(len(data["timeTable"])):
            dataframe3[f"Day {day + 1}"] = ["\n".join(student["subs"][day]) for student in data["timeTableSummary"]]

        dataframe4 = pd.DataFrame([{
            "2 exams in 1 day": str(data["Summary"][0]),
            "3 exams in 2 days": str(data["Summary"][1]),
            "4 exams in 2 days": str(data["Summary"][2]),
        }])

        excel_path = "TimeTable.xlsx"

        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            dataframe1.to_excel(writer, index=False, sheet_name="TimeTable")
            worksheet = writer.sheets["TimeTable"]
            format_sheet(worksheet, dataframe1)

            dataframe4.to_excel(writer, index=False, sheet_name="TimeTable", startrow=len(dataframe1) + 3)

            dataframe5.to_excel(writer, index=False, sheet_name="SlotStrength")
            worksheet = writer.sheets["SlotStrength"]
            format_sheet(worksheet, dataframe5)

            dataframe2.to_excel(writer, index=False, sheet_name="StudentSummary1")
            worksheet = writer.sheets["StudentSummary1"]
            format_sheet(worksheet, dataframe2)

            dataframe3.to_excel(writer, index=False, sheet_name="StudentSummary2")
            worksheet = writer.sheets["StudentSummary2"]
            format_sheet(worksheet, dataframe3)

        print("Schedule is generated.")
    except Exception as e:
        print(f"Error generating Excel: {str(e)}", file=sys.stderr)
        sys.exit(1)


# Run if executed via CLI
if __name__ == "__main__":
    # Read the input data from stdin
    try:
        data = json.loads(sys.stdin.read())  # Read from stdin instead of arguments
        json_to_excel(data)
    except Exception as e:
        print(f"Error reading JSON input: {str(e)}", file=sys.stderr)
        sys.exit(1)
